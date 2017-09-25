#! usr/bin/python
#coding=utf-8

'''
模块名称：
模块主要功能：
模块实现的方法：
模块对外接口：
模块作者：
编写时间：
修改说明：
修改时间：
'''

import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook


def todoCreateExcel():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'ID'
    ws['B1'] = '待办事项'
    ws['C1'] = '开始时间'
    ws['D1'] = '结束时间'
    ws['E1'] = '事项状态'
    wb.save('todoSave.xlsx')

def todoWriteExcel(todoStr, todoStartTime, todoEndTime):
    wb = load_workbook('todoSave.xlsx')
    ws = wb.active
    cur_write_row = ws.max_row + 1
    ws['A'+str(cur_write_row)] = cur_write_row - 1
    ws['B'+str(cur_write_row)] = todoStr
    ws['C'+str(cur_write_row)] = todoStartTime
    ws['D'+str(cur_write_row)] = todoEndTime
    ws['E'+str(cur_write_row)] = '进行中'
    wb.save('todoSave.xlsx')

def todoReadExcel(todoID):
    wb = load_workbook('todoSave.xlsx')
    ws = wb.active
    todoStr = ws['B'+ str(todoID+1)]
    todoStartTime = ws['C'+ str(todoID+1)]
    todoEndTime = ws['D'+ str(todoID+1)]
    todoStatus = ws['E'+ str(todoID+1)]
    return [todoStr, todoStartTime, todoEndTime, todoStatus]

#todo:完成按开始时间进行查询并输出的功能

#todo：完成按状态进行查询并输出的功能

todoCreateExcel()
todoWriteExcel(1, 1, 1, 1)